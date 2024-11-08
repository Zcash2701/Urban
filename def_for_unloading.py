import openpyxl as op
from openpyxl.styles import PatternFill
from datetime import datetime
def by_temperature(file, min_pom, max_pom, min_cont, max_cont):
    book = op.load_workbook(file)
    sheet = book.active
    max_vols = sheet.max_row
    fill_blue = PatternFill('solid', fgColor='93ccdd')
    fill_red = PatternFill('solid', fgColor='e6b8b8')
    fill_purpl = PatternFill('solid', fgColor='cdc0da')

    for i_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=i_id, column=13).value)
            if num == 'SNMP No-Such-Instance':
                continue
            elif num <= int(min_pom):
                sheet[f'M{i_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'M{i_id}'].fill = fill_red
        except:
            continue

    for j_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=j_id, column=14).value)
            if num == 'SNMP No-Such-Instance':
                continue

            elif num <= int(min_pom):
                sheet[f'N{j_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'N{j_id}'].fill = fill_red
        except:
            continue

    for e_id in range(max_vols + 1):
        try:
            num = str(sheet.cell(row=e_id, column=7).value)
            if num == 'Есть':
                sheet[f'G{e_id}'].fill = fill_purpl

        except:
            continue

    for f_id in range(max_vols + 1):
        try:

            num = int(sheet.cell(row=f_id, column=6).value)
            if num == 'SNMP No-Such-Instance':
                continue

            elif num <= (min_cont):
                sheet[f'F{f_id}'].fill = fill_blue
            elif num >= (max_cont):
                sheet[f'F{f_id}'].fill = fill_red
        except:
            continue

    now_time = datetime.now()
    file_name = f'Выгрузка {now_time.day}.{now_time.month}.{now_time.year}.xlsx'
    file_path = f'D://Python//Project//Urban//Выгрузки//{file_name}'
    book.save(file_path)
    return file_name

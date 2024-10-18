import tkinter as tk
from tkinter import filedialog as fd
import openpyxl as op
from openpyxl.styles import PatternFill
from datetime import datetime

filename = ''


def open_file():

    global filename
    filename = str(fd.askopenfilenames())
    filename = filename.replace('/', '//').replace('(', '').replace(')', '').replace("'", '').replace(",", '')
    #print((filename))
    answer_file_name_entry.insert(0, filename)
    return




window = tk.Tk()
window.title('Сплав В.')
window.geometry('500x300')
window.resizable(False, False)

button_add_file = tk.Button(text='Выбрать файл', command=open_file, height=1)
button_add_file.place(x=25, y=50)

label_file_name = tk.Label(text="Выбранный файл: ", height=2)
label_file_name.place(x=130, y=25)
answer_file_name_entry = tk.Entry(width=50)
answer_file_name_entry.place(x=130, y=50)



# Лабель и ввод минимальных помещение
label_get_min_tem_pom = tk.Label(text='Min t пом.: ')
label_get_min_tem_pom.place(x=25, y=85)
entry_min_temp_pom = tk.Entry(width=15)
entry_min_temp_pom.place(x=130, y=85)
entry_min_temp_pom.insert(0, '15')

# Лабель и ввод максимальных помещение
label_get_max_tem_pom = tk.Label(text='Max t пом.: ')
label_get_max_tem_pom.place(x=25, y=110)
entry_max_temp_pom = tk.Entry(width=15)
entry_max_temp_pom.place(x=130, y=110)
entry_max_temp_pom.insert(0, '20')


# Лабель и ввод минимальных контур
label_get_min_temp_cont = tk.Label(text='Min t контура: ')
label_get_min_temp_cont.place(x=250, y=85)
entry_min_temp_cont = tk.Entry(width=15)
entry_min_temp_cont.place(x=340, y=85)
entry_min_temp_cont.insert(0, '20')

## Лабель и ввод максимальных контур
label_get_max_temp_cont = tk.Label(text='Max t контура: ')
label_get_max_temp_cont.place(x=250, y=110)
entry_max_temp_cont = tk.Entry(width=15)
entry_max_temp_cont.place(x=340, y=110)
entry_max_temp_cont.insert(0, '25')



entry_ready_answer = tk.Entry(width=30)
entry_ready_answer.place(x=150, y=150)

def get_ready_file():
    global filename


    min_pom = int(entry_min_temp_pom.get())
    max_pom = int(entry_max_temp_pom.get())
    min_cont = int(entry_min_temp_cont.get())
    max_cont = int(entry_max_temp_cont.get())

    book = op.load_workbook(filename)
    sheet = book.active
    max_vols = sheet.max_row
    fill_blue = PatternFill('solid', fgColor='93ccdd')
    fill_red = PatternFill('solid', fgColor='e6b8b8')
    fill_purpl = PatternFill('solid', fgColor='cdc0da')

    for i_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=i_id, column=13).value)
            if num == 'SNMP No-Such-Instance':
                continue
            elif num <= int(min_pom):
                sheet[f'M{i_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'M{i_id}'].fill = fill_red
        except:
            continue

    for j_id in range(max_vols + 1):
        try:
            num = int(sheet.cell(row=j_id, column=14).value)
            if num == 'SNMP No-Such-Instance':
                continue

            elif num <= int(min_pom):
                sheet[f'N{j_id}'].fill = fill_blue
            elif num >= int(max_pom):
                sheet[f'N{j_id}'].fill = fill_red
        except:
            continue

    for e_id in range(max_vols + 1):
        try:
            num = str(sheet.cell(row=e_id, column=7).value)
            if num == 'Есть':
                sheet[f'G{e_id}'].fill = fill_purpl

        except:
            continue

    for f_id in range(max_vols + 1):
        try:

            num = int(sheet.cell(row=f_id, column=6).value)
            if num == 'SNMP No-Such-Instance':
                continue

            elif num <= (min_cont):
                sheet[f'F{f_id}'].fill = fill_blue
            elif num >= (max_cont):
                sheet[f'F{f_id}'].fill = fill_red
        except:
            continue

    now_time = datetime.now()
    file_name = f'Выгрузка {now_time.day}.{now_time.month}.{now_time.year}.xlsx'
    file_path = f'D://Python//Project//Urban//Выгрузки//{file_name}'
    book.save(file_path)
    entry_ready_answer.insert(0, f'{file_name} готова')


do_it_button = tk.Button(text='Выполнить', command=get_ready_file)
do_it_button.place(x=25, y=150)


window.mainloop()